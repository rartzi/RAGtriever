from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .base import Extracted

@dataclass
class XlsxExtractor:
    supported_suffixes = (".xlsx",)

    def extract(self, path: Path) -> Extracted:
        """Extract spreadsheet content.

        TODO: Implement table-aware extraction; include headers and ranges.
        """
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise RuntimeError("openpyxl required for XLSX extraction") from e

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        out: list[str] = []
        for ws in wb.worksheets:
            out.append(f"[[[SHEET {ws.title}]]]")
            # naive: sample first 50 rows/20 cols
            max_r = min(ws.max_row or 0, 50)
            max_c = min(ws.max_column or 0, 20)
            for r in range(1, max_r + 1):
                row_vals = []
                for c in range(1, max_c + 1):
                    v = ws.cell(row=r, column=c).value
                    row_vals.append("" if v is None else str(v))
                out.append("\t".join(row_vals))
            out.append("")
        text = "\n".join(out).strip()
        meta: dict[str, Any] = {"sheet_count": len(wb.worksheets)}
        return Extracted(text=text, metadata=meta)
